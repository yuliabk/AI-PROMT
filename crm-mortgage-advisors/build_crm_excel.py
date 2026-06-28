"""
בונה קובץ Excel מלא למערכת CRM יועצי משכנתאות.
מריצים: python3 build_crm_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime
import re

OUTPUT_FILE = "CRM_יועצי_משכנתאות.xlsx"

# ─── צבעים ───────────────────────────────────────────────────────────────────
BLUE_DARK    = "1A73E8"
BLUE_LIGHT   = "E8F0FE"
GREEN_DARK   = "137333"
GREEN_LIGHT  = "E6F4EA"
ORANGE_DARK  = "E65100"
ORANGE_LIGHT = "FFF3E0"
YELLOW_DARK  = "594300"
YELLOW_LIGHT = "FCE8B2"
RED_DARK     = "A50E0E"
RED_LIGHT    = "FDE7F3"
PURPLE_DARK  = "6A1B9A"
PURPLE_LIGHT = "F3E5F5"
GREY_DARK    = "37474F"
GREY_LIGHT   = "ECEFF1"
WHITE        = "FFFFFF"
HEADER_BG    = "1A73E8"
ROW_ALT      = "F8F9FA"

# ─── סטטוסים ────────────────────────────────────────────────────────────────
CASE_STATUSES = [
    "ליד חדש",
    "בתהליך איסוף מסמכים",
    "הוגש לאישור",
    "התקבל אישור עקרוני",
    "הוגש לאישור סופי",
    "אושר סופי",
    "הועבר לנוטריון",
    "נחתם - עסקה סגורה",
    "לא סגור כי המפתח אצל הג'ינג'י",
    "כפוף להחלטת בגץ"ת
    "מחכה לאישור של חביתוש"
]

DOC_STATUSES = ["ממתין למסמכים", "חלקי", "תקין מלא"]

BANKS = [
    "לאומי", "הפועלים", "דיסקונט", "מזרחי טפחות",
    "מרכנתיל דיסקונט", "אוצר החייל", "ירושלים", "אגוד",
]

SOURCES = [
    "אתר אינטרנט", "המלצה", "רשתות חברתיות",
    "גוגל", "פרסום", "אחר",
]

STATUS_COLORS = {
    "ליד חדש":                 (BLUE_LIGHT,   BLUE_DARK),
    "בתהליך איסוף מסמכים":    (ORANGE_LIGHT, ORANGE_DARK),
    "הוגש לאישור":             (YELLOW_LIGHT, YELLOW_DARK),
    "התקבל אישור עקרוני":      (GREEN_LIGHT,  GREEN_DARK),
    "הוגש לאישור סופי":        ("D2E3FC",     "1967D2"),
    "אושר סופי":               ("B7E1CD",     "0D652D"),
    "הועבר לנוטריון":          (RED_LIGHT,    RED_DARK),
    "נחתם - עסקה סגורה":       (GREY_DARK,    WHITE),
}

DOC_COLORS = {
    "ממתין למסמכים": (YELLOW_LIGHT, YELLOW_DARK),
    "חלקי":          (ORANGE_LIGHT, ORANGE_DARK),
    "תקין מלא":      (GREEN_LIGHT,  GREEN_DARK),
}


# ─── עזר ────────────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=None, size=11, italic=False):
    kw = dict(bold=bold, size=size, italic=italic)
    if color:
        kw["color"] = color
    return Font(**kw)

def align(h="right", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def header_border():
    b = Side(style="medium", color=HEADER_BG)
    return Border(bottom=b)


# ─── גיליון 1: לקוחות ────────────────────────────────────────────────────────

def build_clients_sheet(wb):
    ws = wb.active
    ws.title = "לקוחות"
    ws.sheet_view.rightToLeft = True

    # ─ כותרות
    headers = [
        ("מזהה לקוח",          "A", 18),
        ("שם מלא",             "B", 20),
        ("טלפון",              "C", 14),
        ("מייל",               "D", 26),
        ("מקור הגעה",          "E", 17),
        ("תאריך הצטרפות",      "F", 17),
        ("בנק מטפל",           "G", 16),
        ("סטטוס תיק",          "H", 22),
        ("סטטוס מסמכים",       "I", 18),
        ("תאריך עדכון אחרון",  "J", 20),
        ("קישור תיקייה",       "K", 14),
        ("קישור פגישה",        "L", 14),
        ("קישור WhatsApp",     "M", 16),
        ("הערות",              "N", 30),
    ]

    for col_idx, (title, col_letter, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill      = fill(HEADER_BG)
        cell.font      = font(bold=True, color=WHITE, size=11)
        cell.alignment = align("center")
        cell.border    = header_border()
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # ─ נתוני דמו
    sample_data = [
        ["CRM-20240601-4827", "ישראל ישראלי",   "0501234567", "israel@example.com",     "המלצה",          datetime.date(2024,6,1),  "מזרחי טפחות",  "התקבל אישור עקרוני",   "תקין מלא",      datetime.date(2024,6,10), "📁", "📅", "💬", "לקוח VIP"],
        ["CRM-20240603-2914", "רחל כהן",         "0521111222", "rachel@example.com",     "אתר אינטרנט",    datetime.date(2024,6,3),  "לאומי",         "בתהליך איסוף מסמכים", "חלקי",          datetime.date(2024,6,8),  "📁", "📅", "💬", ""],
        ["CRM-20240605-7731", "משה לוי",         "0541234000", "moshe@example.com",      "גוגל",           datetime.date(2024,6,5),  "הפועלים",       "הוגש לאישור",          "תקין מלא",      datetime.date(2024,6,9),  "📁", "📅", "💬", ""],
        ["CRM-20240607-3388", "שרה אברהם",       "0509876543", "sara@example.com",       "רשתות חברתיות",  datetime.date(2024,6,7),  "דיסקונט",       "ליד חדש",              "ממתין למסמכים", datetime.date(2024,6,7),  "",   "📅", "💬", ""],
        ["CRM-20240608-9921", "דוד מזרחי",       "0527654321", "david@example.com",      "המלצה",          datetime.date(2024,6,8),  "מזרחי טפחות",  "הוגש לאישור סופי",    "תקין מלא",      datetime.date(2024,6,11), "📁", "📅", "💬", ""],
        ["CRM-20240609-1155", "מרים גולדברג",    "0531122334", "miriam@example.com",     "פרסום",          datetime.date(2024,6,9),  "אוצר החייל",    "נחתם - עסקה סגורה",   "תקין מלא",      datetime.date(2024,6,12), "📁", "📅", "💬", "עסקה ראשונה"],
        ["CRM-20240610-4466", "יעקב שמעון",      "0548888777", "yaakov@example.com",     "אחר",            datetime.date(2024,6,10), "הפועלים",       "אושר סופי",           "תקין מלא",      datetime.date(2024,6,13), "📁", "📅", "💬", ""],
        ["CRM-20240611-6677", "לאה רוזנברג",     "0523344556", "leah@example.com",       "אתר אינטרנט",    datetime.date(2024,6,11), "לאומי",         "הועבר לנוטריון",       "תקין מלא",      datetime.date(2024,6,14), "📁", "📅", "💬", ""],
        ["CRM-20240612-2233", "אבי כץ",          "0545566778", "avi@example.com",        "גוגל",           datetime.date(2024,6,12), "מרכנתיל דיסקונט","ליד חדש",             "ממתין למסמכים", datetime.date(2024,6,12), "",   "📅", "💬", ""],
        ["CRM-20240613-8899", "נעמי שפירא",      "0502233445", "naomi@example.com",      "המלצה",          datetime.date(2024,6,13), "מזרחי טפחות",  "בתהליך איסוף מסמכים","ממתין למסמכים",datetime.date(2024,6,13), "",   "📅", "💬", ""],
    ]

    for r_idx, row in enumerate(sample_data, start=2):
        bg = WHITE if r_idx % 2 == 0 else ROW_ALT
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = thin_border()
            cell.alignment = align("right" if c_idx not in (3,4,6,10,11,12,13) else "center")
            cell.fill      = fill(bg)
            cell.font      = font(size=10)
            if isinstance(val, datetime.date):
                cell.number_format = "DD/MM/YYYY"

        ws.row_dimensions[r_idx].height = 22

    # ─ ולידציות
    dv_status = DataValidation(
        type="list",
        formula1='"' + ','.join(CASE_STATUSES) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_status.error      = "בחר סטטוס תקין מהרשימה"
    dv_status.errorTitle = "סטטוס לא חוקי"
    dv_status.sqref      = "H2:H1000"
    ws.add_data_validation(dv_status)

    dv_doc = DataValidation(
        type="list",
        formula1='"' + ','.join(DOC_STATUSES) + '"',
        allow_blank=True,
    )
    dv_doc.sqref = "I2:I1000"
    ws.add_data_validation(dv_doc)

    dv_bank = DataValidation(
        type="list",
        formula1='"' + ','.join(BANKS) + '"',
        allow_blank=True,
    )
    dv_bank.sqref = "G2:G1000"
    ws.add_data_validation(dv_bank)

    dv_source = DataValidation(
        type="list",
        formula1='"' + ','.join(SOURCES) + '"',
        allow_blank=True,
    )
    dv_source.sqref = "E2:E1000"
    ws.add_data_validation(dv_source)

    # ─ עיצוב מותנה – סטטוס תיק (עמודה H)
    for status, (bg_hex, fg_hex) in STATUS_COLORS.items():
        rule = Rule(
            type="containsText",
            operator="containsText",
            text=status,
            dxf=DifferentialStyle(
                fill=PatternFill(bgColor=bg_hex),
                font=Font(color=fg_hex, bold=True),
            ),
        )
        rule.formula = [f'NOT(ISERROR(SEARCH("{status}",H2)))']
        ws.conditional_formatting.add("H2:H1000", rule)

    # ─ עיצוב מותנה – סטטוס מסמכים (עמודה I)
    for status, (bg_hex, fg_hex) in DOC_COLORS.items():
        rule = Rule(
            type="containsText",
            operator="containsText",
            text=status,
            dxf=DifferentialStyle(
                fill=PatternFill(bgColor=bg_hex),
                font=Font(color=fg_hex, bold=True),
            ),
        )
        rule.formula = [f'NOT(ISERROR(SEARCH("{status}",I2)))']
        ws.conditional_formatting.add("I2:I1000", rule)

    # ─ טבלת Excel
    tab = Table(displayName="TblClients", ref=f"A1:N{len(sample_data)+1}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    # ─ הגדרת הדפסה
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1

    return ws


# ─── גיליון 2: דשבורד ────────────────────────────────────────────────────────

def build_dashboard_sheet(wb):
    ws = wb.create_sheet("דשבורד")
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    def section_title(row, col, text, span=6):
        cell = ws.cell(row=row, column=col, value=text)
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
        cell.fill      = fill(BLUE_LIGHT)
        cell.font      = font(bold=True, color=BLUE_DARK, size=12)
        cell.alignment = align("right")
        ws.row_dimensions[row].height = 28

    def kpi_box(row, col, label, value, bg, fg, span=1):
        lbl = ws.cell(row=row, column=col, value=label)
        lbl.fill      = fill(bg)
        lbl.font      = font(bold=True, color=fg, size=10)
        lbl.alignment = align("center")
        lbl.border    = thin_border()
        ws.row_dimensions[row].height = 22

        val = ws.cell(row=row+1, column=col, value=value)
        val.fill      = fill(bg)
        val.font      = font(bold=True, color=fg, size=28)
        val.alignment = align("center")
        val.border    = thin_border()
        ws.row_dimensions[row+1].height = 50

    def table_header(row, col, headers):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=col+i, value=h)
            c.fill      = fill("F1F3F4")
            c.font      = font(bold=True, size=10)
            c.alignment = align("center")
            c.border    = thin_border()
        ws.row_dimensions[row].height = 22

    def table_row(row, col, values, bg=WHITE):
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=col+i, value=v)
            c.fill      = fill(bg)
            c.font      = font(size=10)
            c.alignment = align("right" if i == 0 else "center")
            c.border    = thin_border()
        ws.row_dimensions[row].height = 20

    # ─ כותרת ראשית
    ws.merge_cells("A1:H1")
    title = ws.cell(row=1, column=1,
                    value="📊  דשבורד ניהול תיקי משכנתאות  |  עודכן: " +
                          datetime.date.today().strftime("%d/%m/%Y"))
    title.fill      = fill(HEADER_BG)
    title.font      = font(bold=True, color=WHITE, size=14)
    title.alignment = align("center")
    ws.row_dimensions[1].height = 40

    # ─ KPI boxes  (שורות 3-4)
    kpis = [
        ("סה\"כ לקוחות",   10, BLUE_LIGHT,   BLUE_DARK),
        ("תיקים פעילים",    9, GREEN_LIGHT,  GREEN_DARK),
        ("עסקאות סגורות",   1, GREY_LIGHT,   GREY_DARK),
        ("מסמכים חסרים",    4, YELLOW_LIGHT, YELLOW_DARK),
        ("קרובים לחתימה",   2, PURPLE_LIGHT, PURPLE_DARK),
    ]
    for i, (label, value, bg, fg) in enumerate(kpis):
        kpi_box(3, i+1, label, value, bg, fg)

    # ─ סעיף 1: תיקים לפי סטטוס  (שורה 6+)
    section_title(6, 1, "📋  תיקים לפי סטטוס", 4)
    table_header(7, 1, ["סטטוס תיק", "תיקים", "%"])

    status_data = [
        ("ליד חדש",               2),
        ("בתהליך איסוף מסמכים",   2),
        ("הוגש לאישור",            1),
        ("התקבל אישור עקרוני",     1),
        ("הוגש לאישור סופי",       1),
        ("אושר סופי",              1),
        ("הועבר לנוטריון",         1),
        ("נחתם - עסקה סגורה",      1),
    ]
    total = sum(v for _, v in status_data)
    for i, (s, v) in enumerate(status_data):
        bg_hex, fg_hex = STATUS_COLORS.get(s, (WHITE, "000000"))
        row = 8 + i
        ws.cell(row=row, column=1, value=s).fill    = fill(bg_hex)
        ws.cell(row=row, column=1).font             = font(color=fg_hex, size=10, bold=True)
        ws.cell(row=row, column=1).alignment        = align("right")
        ws.cell(row=row, column=1).border           = thin_border()
        ws.cell(row=row, column=2, value=v).fill    = fill(bg_hex)
        ws.cell(row=row, column=2).font             = font(color=fg_hex, size=10, bold=True)
        ws.cell(row=row, column=2).alignment        = align("center")
        ws.cell(row=row, column=2).border           = thin_border()
        pct = f"{round(v/total*100)}%"
        ws.cell(row=row, column=3, value=pct).fill  = fill(bg_hex)
        ws.cell(row=row, column=3).font             = font(color=fg_hex, size=10)
        ws.cell(row=row, column=3).alignment        = align("center")
        ws.cell(row=row, column=3).border           = thin_border()
        ws.row_dimensions[row].height = 22

    # ─ סעיף 2: תיקים לפי בנק (ליד סעיף 1, עמודה 5)
    section_title(6, 5, "🏦  תיקים לפי בנק", 3)
    table_header(7, 5, ["בנק מטפל", "תיקים"])
    bank_data = [
        ("מזרחי טפחות",       3),
        ("לאומי",             2),
        ("הפועלים",           2),
        ("דיסקונט",           1),
        ("מרכנתיל דיסקונט",   1),
        ("אוצר החייל",        1),
    ]
    for i, (b, v) in enumerate(bank_data):
        row = 8 + i
        bg = WHITE if i % 2 == 0 else ROW_ALT
        table_row(row, 5, [b, v], bg)

    # ─ סעיף 3: סטטוס מסמכים (שורה 17+)
    section_title(17, 1, "📄  סטטוס מסמכים", 4)
    table_header(18, 1, ["סטטוס מסמכים", "תיקים", "%"])
    doc_data = [("ממתין למסמכים", 3), ("חלקי", 2), ("תקין מלא", 5)]
    doc_total = sum(v for _, v in doc_data)
    for i, (s, v) in enumerate(doc_data):
        bg_hex, fg_hex = DOC_COLORS.get(s, (WHITE, "000000"))
        row = 19 + i
        ws.cell(row=row, column=1, value=s).fill    = fill(bg_hex)
        ws.cell(row=row, column=1).font             = font(color=fg_hex, bold=True, size=10)
        ws.cell(row=row, column=1).alignment        = align("right")
        ws.cell(row=row, column=1).border           = thin_border()
        ws.cell(row=row, column=2, value=v).fill    = fill(bg_hex)
        ws.cell(row=row, column=2).font             = font(color=fg_hex, bold=True, size=10)
        ws.cell(row=row, column=2).alignment        = align("center")
        ws.cell(row=row, column=2).border           = thin_border()
        ws.cell(row=row, column=3, value=f"{round(v/doc_total*100)}%").fill = fill(bg_hex)
        ws.cell(row=row, column=3).font             = font(color=fg_hex, size=10)
        ws.cell(row=row, column=3).alignment        = align("center")
        ws.cell(row=row, column=3).border           = thin_border()
        ws.row_dimensions[row].height = 22

    # ─ סעיף 4: פעילות אחרונה (שורה 24+)
    section_title(24, 1, "🕐  10 הלקוחות האחרונים שעודכנו", 7)
    table_header(25, 1, ["שם לקוח", "סטטוס תיק", "סטטוס מסמכים", "עדכון אחרון"])
    recent = [
        ("לאה רוזנברג",   "הועבר לנוטריון",        "תקין מלא",      "14/06/2024"),
        ("יעקב שמעון",    "אושר סופי",              "תקין מלא",      "13/06/2024"),
        ("מרים גולדברג",  "נחתם - עסקה סגורה",      "תקין מלא",      "12/06/2024"),
        ("דוד מזרחי",     "הוגש לאישור סופי",       "תקין מלא",      "11/06/2024"),
        ("ישראל ישראלי",  "התקבל אישור עקרוני",     "תקין מלא",      "10/06/2024"),
        ("רחל כהן",       "בתהליך איסוף מסמכים",   "חלקי",          "08/06/2024"),
        ("משה לוי",       "הוגש לאישור",             "תקין מלא",      "09/06/2024"),
        ("שרה אברהם",     "ליד חדש",                "ממתין למסמכים", "07/06/2024"),
    ]
    for i, r in enumerate(recent):
        bg = WHITE if i % 2 == 0 else ROW_ALT
        table_row(26+i, 1, list(r), bg)

    # ─ גרף עמודות – תיקים לפי סטטוס
    chart_data_start = 8
    chart = BarChart()
    chart.type        = "bar"
    chart.grouping    = "clustered"
    chart.title       = "תיקים לפי סטטוס"
    chart.y_axis.title = "מספר תיקים"
    chart.style       = 10
    chart.width       = 14
    chart.height      = 10

    data   = Reference(ws, min_col=2, min_row=7, max_row=7+len(status_data))
    cats   = Reference(ws, min_col=1, min_row=8, max_row=7+len(status_data))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E17")

    # ─ רוחב עמודות
    for col, w in [("A",24),("B",10),("C",8),("D",8),("E",20),("F",10),("G",8),("H",8)]:
        ws.column_dimensions[col].width = w

    return ws


# ─── גיליון 3: לוג פעולות ────────────────────────────────────────────────────

def build_log_sheet(wb):
    ws = wb.create_sheet("לוג פעולות")
    ws.sheet_view.rightToLeft = True

    headers = ["תאריך ושעה", "פעולה", "שם לקוח", "פרטים"]
    widths  = [20, 20, 20, 50]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill      = fill(HEADER_BG)
        cell.font      = font(bold=True, color=WHITE)
        cell.alignment = align("center")
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    log_data = [
        (datetime.datetime(2024,6,1,9,0),  "SETUP",           "",              "התקנה ראשונית הושלמה"),
        (datetime.datetime(2024,6,1,9,5),  "ONBOARDING",      "ישראל ישראלי",  "לקוח חדש נקלט – CRM-20240601-4827"),
        (datetime.datetime(2024,6,1,9,6),  "ONBOARDING",      "רחל כהן",       "לקוח חדש נקלט – CRM-20240603-2914"),
        (datetime.datetime(2024,6,3,10,0), "FOLDER_CREATED",  "ישראל ישראלי",  "https://drive.google.com/..."),
        (datetime.datetime(2024,6,5,14,0), "STATUS_CHANGE",   "משה לוי",       "ליד חדש → הוגש לאישור"),
        (datetime.datetime(2024,6,5,14,1), "STATUS_EMAIL_SENT","משה לוי",      "הוגש לאישור"),
        (datetime.datetime(2024,6,8,9,0),  "REMINDER_RUN",    "",              "סה\"כ תזכורות: נשלחו=3, שגיאות=0"),
        (datetime.datetime(2024,6,10,11,0),"STATUS_CHANGE",   "ישראל ישראלי", "הוגש לאישור → התקבל אישור עקרוני"),
        (datetime.datetime(2024,6,12,9,0), "REMINDER_RUN",    "",              "סה\"כ תזכורות: נשלחו=2, שגיאות=0"),
        (datetime.datetime(2024,6,14,16,0),"STATUS_CHANGE",   "לאה רוזנברג",  "אושר סופי → הועבר לנוטריון"),
    ]

    action_colors = {
        "SETUP":            (BLUE_LIGHT,   BLUE_DARK),
        "ONBOARDING":       (GREEN_LIGHT,  GREEN_DARK),
        "FOLDER_CREATED":   ("EDE7F6",     PURPLE_DARK),
        "STATUS_CHANGE":    (YELLOW_LIGHT, YELLOW_DARK),
        "STATUS_EMAIL_SENT":(GREEN_LIGHT,  GREEN_DARK),
        "REMINDER_RUN":     (ORANGE_LIGHT, ORANGE_DARK),
    }

    for r_idx, (dt, action, client, details) in enumerate(log_data, start=2):
        bg_hex, fg_hex = action_colors.get(action, (WHITE, "000000"))
        row_data = [dt, action, client, details]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = thin_border()
            cell.alignment = align("right" if c_idx > 1 else "center")
            cell.font      = font(size=10)
            if isinstance(val, datetime.datetime):
                cell.number_format = "DD/MM/YYYY HH:MM"
            if c_idx == 2:  # עמודת פעולה – צבע רקע
                cell.fill = fill(bg_hex)
                cell.font = font(color=fg_hex, bold=True, size=10)
        ws.row_dimensions[r_idx].height = 20

    return ws


# ─── גיליון 4: מדריך הגדרות ─────────────────────────────────────────────────

def build_guide_sheet(wb):
    ws = wb.create_sheet("מדריך והגדרות")
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    def heading(row, text, level=1):
        cell = ws.cell(row=row, column=1, value=text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        if level == 1:
            cell.fill   = fill(HEADER_BG)
            cell.font   = font(bold=True, color=WHITE, size=13)
            ws.row_dimensions[row].height = 36
        else:
            cell.fill   = fill(BLUE_LIGHT)
            cell.font   = font(bold=True, color=BLUE_DARK, size=11)
            ws.row_dimensions[row].height = 26
        cell.alignment = align("right")

    def body(row, text, indent=False):
        cell = ws.cell(row=row, column=1, value=("    " if indent else "") + text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell.font      = font(size=10)
        cell.alignment = align("right", wrap=True)
        ws.row_dimensions[row].height = 18

    heading(1,  "📖  מדריך שימוש – מערכת CRM יועצי משכנתאות")
    body(2, "")
    heading(3,  "תהליך 1: קליטת לקוח חדש", 2)
    body(4,  "1. הזן שם, טלפון ומייל בגיליון 'לקוחות'.")
    body(5,  "2. מזהה לקוח נוצר אוטומטית ע\"י הסקריפט (CRM-YYYYMMDD-XXXX).")
    body(6,  "3. לחץ על תפריט ⚙️ ← 'פתח תיקייה ללקוח נבחר' לפתיחת תיקייה ב-Drive.")
    body(7,  "4. הודעת WhatsApp ברכה נשלחת אוטומטית דרך Make/n8n.")
    body(8,  "")
    heading(9,  "תהליך 2: ניהול מסמכים", 2)
    body(10, "1. עדכן 'סטטוס מסמכים' לכל לקוח (ממתין / חלקי / תקין מלא).")
    body(11, "2. תזכורות נשלחות אוטומטית בימים ראשון ורביעי לכל לקוח עם חוסרים.")
    body(12, "")
    heading(13, "תהליך 3: מחזור חיי עסקה", 2)
    body(14, "1. שנה 'סטטוס תיק' בעמודה H.")
    body(15, "2. תאריך עדכון אחרון מתעדכן אוטומטית.")
    body(16, "3. מייל HTML מעוצב נשלח ללקוח עם הסבר השלב.")
    body(17, "")
    heading(18, "תהליך 4: תקשורת מהירה", 2)
    body(19, "• לחץ על 💬 בעמודה M לפתיחת WhatsApp Web ישירות לאיש קשר.")
    body(20, "• לחץ על 📅 בעמודה L לפתיחת אירוע ב-Google Calendar עם שם הלקוח.")
    body(21, "")
    heading(22, "תהליך 5: דשבורד שבועי", 2)
    body(23, "• גיליון 'דשבורד' מציג KPIs, חתך לפי סטטוס ובנק, ופעילות אחרונה.")
    body(24, "• לרענון ידני: תפריט ⚙️ ← 'רענן דשבורד'.")
    body(25, "• רענון אוטומטי: כל ראשון בשעה 08:00.")
    body(26, "")
    heading(27, "🔧  הגדרות Webhook (Make / n8n)", 2)
    body(28, "עדכן את WEBHOOK_URL ב-Config.gs:")
    body(29, "    WEBHOOK_URL: 'https://hook.make.com/YOUR_HOOK_ID'", True)
    body(30, "")
    body(31, "אירועים נשלחים ל-webhook:")
    body(32, "    • NEW_CLIENT – לקוח חדש נקלט", True)
    body(33, "    • STATUS_CHANGE – שינוי סטטוס תיק", True)
    body(34, "    • DOCUMENT_REMINDER – תזכורת מסמכים יומית", True)
    body(35, "")
    heading(36, "📋  רשימת סטטוסים חוקיים", 2)

    status_table_headers = ["סטטוס תיק", "צבע", "משמעות"]
    for i, h in enumerate(status_table_headers, start=1):
        c = ws.cell(row=37, column=i, value=h)
        c.fill      = fill("F1F3F4")
        c.font      = font(bold=True, size=10)
        c.alignment = align("center")
        c.border    = thin_border()

    meanings = {
        "ליד חדש":               "פנייה ראשונית – טרם נפתח תיק",
        "בתהליך איסוף מסמכים":  "ממתין למסמכים מהלקוח",
        "הוגש לאישור":           "הבקשה הוגשה לבנק",
        "התקבל אישור עקרוני":    "הבנק אישר עקרונית",
        "הוגש לאישור סופי":      "כל מסמכים הוגשו לאישור סופי",
        "אושר סופי":             "הבנק אישר סופית",
        "הועבר לנוטריון":        "ממתין לחתימה",
        "נחתם - עסקה סגורה":     "עסקה הושלמה",
    }
    for i, (s, m) in enumerate(meanings.items()):
        row = 38 + i
        bg, fg = STATUS_COLORS[s]
        ws.cell(row=row, column=1, value=s).fill      = fill(bg)
        ws.cell(row=row, column=1).font               = font(color=fg, bold=True, size=10)
        ws.cell(row=row, column=1).alignment          = align("right")
        ws.cell(row=row, column=1).border             = thin_border()
        ws.cell(row=row, column=2, value="■").fill    = fill(bg)
        ws.cell(row=row, column=2).font               = font(color=fg, size=14)
        ws.cell(row=row, column=2).alignment          = align("center")
        ws.cell(row=row, column=2).border             = thin_border()
        ws.cell(row=row, column=3, value=m).font      = font(size=10)
        ws.cell(row=row, column=3).alignment          = align("right")
        ws.cell(row=row, column=3).border             = thin_border()
        ws.row_dimensions[row].height = 20

    for col, w in [("A",28),("B",8),("C",35),("D",10),("E",10)]:
        ws.column_dimensions[col].width = w

    return ws


# ─── הרצה ראשית ──────────────────────────────────────────────────────────────

def main():
    wb = Workbook()

    build_clients_sheet(wb)
    build_dashboard_sheet(wb)
    build_log_sheet(wb)
    build_guide_sheet(wb)

    # סדר גיליונות
    for i, name in enumerate(["לקוחות", "דשבורד", "לוג פעולות", "מדריך והגדרות"]):
        sheet = wb[name]
        wb.move_sheet(sheet, offset=i - wb.index(sheet))

    wb.save(OUTPUT_FILE)
    print(f"✅ הקובץ נוצר: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
